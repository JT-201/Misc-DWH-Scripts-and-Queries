import pandas as pd
import mysql.connector
from config import get_db_config
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def get_connection():
    cfg = get_db_config()
    return mysql.connector.connect(**cfg)


query = """
SELECT
    BIN_TO_UUID(s.user_id) AS user_id,
    ap.token AS customer_io_id,
    COUNT(DISTINCT a.id) AS completed_visits
FROM subscription_subscriptions s
JOIN partner_employments pe
    ON pe.user_id = s.user_id
    AND pe.partner_id = UUID_TO_BIN('b0c1f8d2-3e4a-4c5b-9f6d-7e8f9a0b1c2d')
JOIN appointments_appointments a
    ON a.user_id = s.user_id
    AND a.status = 'COMPLETED'
JOIN analytics_profiles ap
    ON ap.user_id = s.user_id
WHERE s.cancellation_date IS NULL
GROUP BY s.user_id, ap.token
ORDER BY completed_visits DESC;
"""

print("Running query...")
conn = get_connection()
cursor = conn.cursor(dictionary=True)
cursor.execute(query)
rows = cursor.fetchall()
cursor.close()
conn.close()

df = pd.DataFrame(rows)
print(f"Pulled {len(df):,} rows")

output_path = '/Users/meganriddle/Documents/9am_dev_queries/thermo_fisher_active_completed_visit.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Contact List')
    ws = writer.sheets['Contact List']
    header_fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_font = Font(name='Arial', size=9)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = row_font
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 14), 45)
    ws.freeze_panes = 'A2'

print(f"Saved to {output_path}")