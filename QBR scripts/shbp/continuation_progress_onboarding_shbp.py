DB_ENV=prod ~/.venv/bin/python continuation_member_pull.py

import pandas as pd
import mysql.connector
from config import get_db_config
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def get_connection():
    cfg = get_db_config()
    return mysql.connector.connect(**cfg)

query = """
WITH
active_members AS (
    SELECT
        u.id                                                AS member_id,
        u.readable_id,
        u.primary_condition_group,
        s.subscription_start_date,
        DATEDIFF(CURDATE(), s.subscription_start_date)      AS days_enrolled
    FROM users u
    JOIN (
        SELECT user_id,
               MIN(start_date)        AS subscription_start_date,
               MAX(cancellation_date) AS cancellation_date
        FROM subscriptions
        WHERE status = 'ACTIVE' AND cancellation_date IS NULL
        GROUP BY user_id
    ) s ON s.user_id = u.id
    WHERE u.id IN (
        SELECT user_id FROM partner_employers WHERE name = 'State of Georgia'
    )
),
questionnaire_answers AS (
    SELECT
        member_id,
        MAX(CASE WHEN question_id = 'A8z9j98E0sxR' AND rn = 1 THEN answer_value END) AS past_glp1_use,
        MAX(CASE WHEN question_id = 'gV9Xu8RzF9hR' AND rn = 1 THEN answer_value END) AS wants_to_continue_glp1,
        MAX(CASE WHEN question_id = 'knzp0ZppEBF4' AND rn = 1 THEN answer_text  END) AS reported_medication_name
    FROM (
        SELECT
            user_id      AS member_id,
            question_id,
            answer_value,
            answer_text,
            ROW_NUMBER() OVER (
                PARTITION BY user_id, question_id ORDER BY answered_at DESC
            ) AS rn
        FROM questionnaire_records
        WHERE question_id IN ('A8z9j98E0sxR', 'gV9Xu8RzF9hR', 'knzp0ZppEBF4')
          AND is_latest_answer = 1
    ) r
    GROUP BY member_id
),
wm_rx_2026 AS (
    SELECT
        p.patient_user_id                                   AS member_id,
        1                                                   AS has_wm_rx_since_2026,
        MIN(DATE(p.prescribed_at))                          AS first_wm_rx_date,
        MAX(DATE(p.prescribed_at))                          AS latest_wm_rx_date,
        MAX(p.prescribed_at)                                AS latest_wm_prescribed_at,
        MAX(m.name)                                         AS latest_wm_rx_drug_name
    FROM prescriptions p
    JOIN medication_dosage_ndcs mdn ON mdn.ndc = p.prescribed_ndc
    JOIN medication_dosages md      ON md.id = mdn.medication_dosage_id
    JOIN medications m              ON m.id = md.medication_id
    WHERE m.therapy_type = 'WM'
      AND DATE(p.prescribed_at) >= '2026-01-01'
    GROUP BY p.patient_user_id
),
first_glp1_rx AS (
    SELECT
        patient_user_id                                     AS member_id,
        MIN(prescribed_at)                                  AS first_glp1_rx_date,
        MAX(CASE WHEN rn = 1 THEN drug_name END)            AS first_glp1_rx_drug_name
    FROM (
        SELECT
            p.patient_user_id,
            p.prescribed_at,
            m.name                                          AS drug_name,
            ROW_NUMBER() OVER (
                PARTITION BY p.patient_user_id ORDER BY p.prescribed_at ASC
            ) AS rn
        FROM prescriptions p
        JOIN medication_dosage_ndcs mdn ON mdn.ndc = p.prescribed_ndc
        JOIN medication_dosages md      ON md.id = mdn.medication_dosage_id
        JOIN medications m              ON m.id = md.medication_id
        JOIN medication_drug_classes mdc ON mdc.medication_id = m.id
        WHERE mdc.drug_class_name = 'GLP1'
    ) ranked
    GROUP BY patient_user_id
),
zendesk_continuation AS (
    SELECT
        zum.user_id                                         AS member_id,
        MIN(zt.created_at)                                  AS continuation_ticket_created_at,
        MIN(CASE WHEN ztt.tag = 'glp1-continuation:ready-for-review'
            THEN zt.created_at END)                         AS ready_for_review_at,
        MIN(CASE WHEN ztt.tag = 'glp1-continuation:assistance-needed'
            THEN zt.created_at END)                         AS assistance_needed_at,
        MAX(CASE WHEN ztt.tag = 'glp1-continuation:ready-for-review'
            THEN zt.status END)                             AS ready_for_review_ticket_status,
        MAX(CASE WHEN ztt.tag = 'glp1-continuation:assistance-needed'
            THEN zt.status END)                             AS assistance_needed_ticket_status,
        CASE
            WHEN MIN(CASE WHEN ztt.tag = 'glp1-continuation:ready-for-review'
                THEN zt.created_at END) IS NOT NULL
                THEN 'ready-for-review'
            WHEN MIN(CASE WHEN ztt.tag = 'glp1-continuation:assistance-needed'
                THEN zt.created_at END) IS NOT NULL
                THEN 'assistance-needed'
            ELSE NULL
        END                                                 AS ticket_type
    FROM zendesk_tickets zt
    JOIN zendesk_ticket_tags ztt    ON ztt.ticket_id = zt.id
    JOIN zendesk_user_mappings zum  ON zum.zendesk_user_id = zt.requester_id
    WHERE ztt.tag IN (
        'glp1-continuation:ready-for-review',
        'glp1-continuation:assistance-needed'
    )
      AND zt.created_at >= '2026-01-01'
    GROUP BY zum.user_id
),
tasks_pivoted AS (
    SELECT
        user_id                                             AS member_id,
        MAX(CASE WHEN slug = 'complete-initial-lab-order'        AND rn = 1 THEN status       END) AS lab_order_status,
        MAX(CASE WHEN slug = 'complete-initial-lab-order'        AND rn = 1 THEN completed_at END) AS lab_order_completed_at,
        MAX(CASE WHEN slug = 'glp1-continuation-questionnaire'   AND rn = 1 THEN status       END) AS questionnaire_status,
        MAX(CASE WHEN slug = 'glp1-continuation-questionnaire'   AND rn = 1 THEN completed_at END) AS questionnaire_completed_at,
        MAX(CASE WHEN slug = 'upload-prescription-label'         AND rn = 1 THEN status       END) AS rx_label_status,
        MAX(CASE WHEN slug = 'upload-prescription-label'         AND rn = 1 THEN completed_at END) AS rx_label_completed_at,
        MAX(CASE WHEN slug = 'pharmacy-insurance'                AND rn = 1 THEN status       END) AS pharmacy_insurance_status,
        MAX(CASE WHEN slug = 'pharmacy-insurance'                AND rn = 1 THEN completed_at END) AS pharmacy_insurance_completed_at,
        MAX(CASE WHEN slug = 'upload-proof-of-weight'            AND rn = 1 THEN status       END) AS proof_of_weight_status,
        MAX(CASE WHEN slug = 'upload-proof-of-weight'            AND rn = 1 THEN completed_at END) AS proof_of_weight_completed_at,
        MAX(CASE WHEN slug = 'preferred-pharmacy'                AND rn = 1 THEN status       END) AS preferred_pharmacy_status,
        MAX(CASE WHEN slug = 'preferred-pharmacy'                AND rn = 1 THEN completed_at END) AS preferred_pharmacy_completed_at
    FROM (
        SELECT user_id, slug, status, completed_at,
               ROW_NUMBER() OVER (
                   PARTITION BY user_id, slug ORDER BY updated_at DESC
               ) AS rn
        FROM tasks
        WHERE slug IN (
            'complete-initial-lab-order',
            'glp1-continuation-questionnaire',
            'upload-prescription-label',
            'pharmacy-insurance',
            'upload-proof-of-weight',
            'preferred-pharmacy'
        )
    ) ranked
    GROUP BY user_id
),
appointment_df AS (
    SELECT
        user_id                                             AS member_id,
        COUNT(*)                                            AS appointments_completed,
        MAX(start)                                          AS last_appointment_session_at
    FROM appointments
    WHERE status = 'COMPLETED'
    GROUP BY user_id
),
last_message AS (
    SELECT
        ae.user_id                                          AS member_id,
        MAX(ae.created_at)                                  AS last_message_sent_at,
        COUNT(*)                                            AS total_messages_sent
    FROM analytics_events ae
    JOIN active_members am ON am.member_id = ae.user_id
    WHERE ae.event_name = 'Message Sent'
    GROUP BY ae.user_id
)
SELECT
    am.readable_id,
    am.primary_condition_group,
    am.subscription_start_date,
    am.days_enrolled,
    NOW()                                                   AS data_as_of,
    COALESCE(qa.past_glp1_use, 0)                           AS is_continuation_member,
    CASE
        WHEN COALESCE(qa.past_glp1_use, 0) = 1 THEN 'Continuation'
        ELSE 'New to Med'
    END                                                     AS member_type,
    COALESCE(qa.wants_to_continue_glp1, 0)                  AS wants_to_continue_glp1,
    qa.reported_medication_name,
    COALESCE(wm.has_wm_rx_since_2026, 0)                    AS has_wm_rx_since_2026,
    wm.first_wm_rx_date,
    wm.latest_wm_rx_date,
    wm.latest_wm_prescribed_at,
    wm.latest_wm_rx_drug_name,
    frx.first_glp1_rx_date,
    frx.first_glp1_rx_drug_name,
    zc.continuation_ticket_created_at,
    zc.ready_for_review_at,
    zc.assistance_needed_at,
    zc.ticket_type,
    zc.ready_for_review_ticket_status,
    zc.assistance_needed_ticket_status,
    tp.questionnaire_status,
    tp.questionnaire_completed_at,
    tp.lab_order_status,
    tp.lab_order_completed_at,
    tp.rx_label_status,
    tp.rx_label_completed_at,
    tp.pharmacy_insurance_status,
    tp.pharmacy_insurance_completed_at,
    tp.proof_of_weight_status,
    tp.proof_of_weight_completed_at,
    tp.preferred_pharmacy_status,
    tp.preferred_pharmacy_completed_at,
    CASE WHEN tp.questionnaire_status = 'COMPLETED' THEN 1 ELSE 0 END
                                                            AS quest_completed,
    CASE WHEN tp.lab_order_status IN ('COMPLETED','SKIPPED') THEN 1 ELSE 0 END
                                                            AS lab_completed,
    CASE
        WHEN tp.lab_order_status IN ('COMPLETED','SKIPPED')
         AND tp.questionnaire_status = 'COMPLETED'
        THEN GREATEST(
            COALESCE(tp.lab_order_completed_at, '1900-01-01'),
            COALESCE(tp.questionnaire_completed_at, '1900-01-01')
        )
        ELSE NULL
    END                                                     AS both_tasks_completed_at,
    CASE
        WHEN COALESCE(wm.has_wm_rx_since_2026, 0) = 1
            THEN 'Has WM Rx'
        WHEN tp.lab_order_status IN ('COMPLETED','SKIPPED')
         AND tp.questionnaire_status = 'COMPLETED'
            THEN 'Ready for Rx'
        WHEN tp.lab_order_status IN ('COMPLETED','SKIPPED')
         AND COALESCE(tp.questionnaire_status,'') != 'COMPLETED'
            THEN 'Missing Questionnaire'
        WHEN COALESCE(tp.lab_order_status,'') NOT IN ('COMPLETED','SKIPPED')
         AND tp.questionnaire_status = 'COMPLETED'
            THEN 'Missing Lab'
        ELSE 'Missing Both'
    END                                                     AS member_category,
    CASE
        WHEN zc.continuation_ticket_created_at IS NOT NULL
         AND tp.questionnaire_completed_at IS NOT NULL
        THEN GREATEST(
            zc.continuation_ticket_created_at,
            tp.questionnaire_completed_at
        )
        WHEN zc.continuation_ticket_created_at IS NOT NULL
        THEN zc.continuation_ticket_created_at
        ELSE NULL
    END                                                     AS clock_start,
    CASE
        WHEN zc.continuation_ticket_created_at IS NOT NULL
         AND tp.questionnaire_completed_at IS NOT NULL
         AND zc.continuation_ticket_created_at < tp.questionnaire_completed_at
            THEN 1
        ELSE 0
    END                                                     AS questionnaire_used_as_clock_start,
    CASE
        WHEN zc.continuation_ticket_created_at IS NOT NULL
         AND tp.questionnaire_completed_at IS NOT NULL
            THEN 1
        ELSE 0
    END                                                     AS has_usable_clock_start,
    CASE
        WHEN wm.first_wm_rx_date IS NOT NULL
         AND zc.continuation_ticket_created_at IS NOT NULL
         AND tp.questionnaire_completed_at IS NOT NULL
        THEN DATEDIFF(
            wm.first_wm_rx_date,
            DATE(GREATEST(
                zc.continuation_ticket_created_at,
                tp.questionnaire_completed_at
            ))
        )
        ELSE NULL
    END                                                     AS days_clock_start_to_rx,
    CASE
        WHEN wm.latest_wm_prescribed_at IS NOT NULL
         AND zc.continuation_ticket_created_at IS NOT NULL
         AND tp.questionnaire_completed_at IS NOT NULL
        THEN ROUND(
            TIMESTAMPDIFF(HOUR,
                GREATEST(
                    zc.continuation_ticket_created_at,
                    tp.questionnaire_completed_at
                ),
                wm.latest_wm_prescribed_at
            ) / 24.0 * (5.0 / 7.0),
            1
        )
        ELSE NULL
    END                                                     AS approx_business_days_to_rx,
    DATEDIFF(CURDATE(), am.subscription_start_date)         AS days_since_enrollment,
    CASE
        WHEN wm.first_wm_rx_date IS NOT NULL
        THEN DATEDIFF(wm.first_wm_rx_date, am.subscription_start_date)
        ELSE NULL
    END                                                     AS days_enrollment_to_rx,
    CASE
        WHEN wm.latest_wm_prescribed_at IS NOT NULL
         AND tp.lab_order_status IN ('COMPLETED','SKIPPED')
         AND tp.questionnaire_status = 'COMPLETED'
        THEN DATEDIFF(
            DATE(wm.latest_wm_prescribed_at),
            DATE(GREATEST(
                COALESCE(tp.lab_order_completed_at, '1900-01-01'),
                COALESCE(tp.questionnaire_completed_at, '1900-01-01')
            ))
        )
        ELSE NULL
    END                                                     AS days_tasks_to_rx,
    CASE
        WHEN COALESCE(wm.has_wm_rx_since_2026, 0) = 0
         AND tp.lab_order_status IN ('COMPLETED','SKIPPED')
         AND tp.questionnaire_status = 'COMPLETED'
        THEN DATEDIFF(CURDATE(), DATE(GREATEST(
            COALESCE(tp.lab_order_completed_at, '1900-01-01'),
            COALESCE(tp.questionnaire_completed_at, '1900-01-01')
        )))
        ELSE NULL
    END                                                     AS days_waiting_since_ready,
    co.appointments_completed,
    co.last_appointment_session_at,
    lm.last_message_sent_at,
    lm.total_messages_sent,
    DATEDIFF(CURDATE(), DATE(lm.last_message_sent_at))      AS days_since_last_message
FROM active_members am
LEFT JOIN questionnaire_answers qa  ON qa.member_id  = am.member_id
LEFT JOIN wm_rx_2026            wm  ON wm.member_id  = am.member_id
LEFT JOIN first_glp1_rx         frx ON frx.member_id = am.member_id
LEFT JOIN zendesk_continuation  zc  ON zc.member_id  = am.member_id
LEFT JOIN tasks_pivoted         tp  ON tp.member_id  = am.member_id
LEFT JOIN appointment_df        co  ON co.member_id  = am.member_id
LEFT JOIN last_message          lm  ON lm.member_id  = am.member_id
ORDER BY am.subscription_start_date DESC
"""
print("Running query...")
conn = get_connection()
df = pd.read_sql(query, conn)
conn.close()
print(f"Pulled {len(df):,} rows, {len(df.columns)} columns")

output_path = 'continuation_members_all_tasks.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
    df.to_excel(writer, index=False, sheet_name='Continuation Members')
    ws = writer.sheets['Continuation Members']

    # Header styling
    header_fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row font
    row_font = Font(name='Arial', size=9)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = row_font

    # Auto column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 35)
    # Freeze header row
    ws.freeze_panes = 'A2'
    
print(f"Saved to {output_path}")
